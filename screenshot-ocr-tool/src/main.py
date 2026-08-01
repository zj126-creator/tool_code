# -*- coding: utf-8 -*-
"""
截图OCR工具 - 桌面应用
功能：截图识别文字、表格，支持导出Excel/CSV/Markdown，支持翻译
OCR引擎：腾讯云OCR API
"""

import sys
import os
import json
import base64
import time
import hashlib
import hmac
import random
import string
import datetime
import csv
import io

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QTextEdit, QTabWidget, QTableWidget,
    QTableWidgetItem, QFileDialog, QMessageBox, QSystemTrayIcon,
    QMenu, QAction, QInputDialog, QComboBox, QGroupBox, QSplitter,
    QStatusBar, QCheckBox, QProgressBar
)
from PyQt5.QtCore import Qt, pyqtSignal, QObject, QTimer, QRect, QPoint
from PyQt5.QtGui import QIcon, QPixmap, QImage, QPainter, QPen, QColor, QFont, QClipboard

import requests
from PIL import Image
import mss
import mss.tools

# ============================================================
# 腾讯云 OCR 配置
# ============================================================

CONFIG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.json")

DEFAULT_CONFIG = {
    "secret_id": "",
    "secret_key": "",
    "region": "ap-guangzhou",
    "translate_enabled": False,
    "translate_lang": "en",
}

def load_config():
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                cfg = json.load(f)
            # 合并默认值
            for k, v in DEFAULT_CONFIG.items():
                if k not in cfg:
                    cfg[k] = v
            return cfg
        except:
            pass
    return DEFAULT_CONFIG.copy()

def save_config(cfg):
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)


# ============================================================
# 腾讯云 API 签名工具
# ============================================================

class TencentCloudSigner:
    """腾讯云 API v3 签名实现"""

    @staticmethod
    def sign(secret_id, secret_key, service, params, action, region, version="2018-11-19"):
        """生成签名并发送请求"""
        host = f"{service}.tencentcloudapi.com"
        endpoint = f"https://{host}"

        timestamp = int(time.time())
        date = datetime.datetime.utcfromtimestamp(timestamp).strftime("%Y-%m-%d")

        # 1. 拼接规范请求串
        http_request_method = "POST"
        canonical_uri = "/"
        canonical_querystring = ""
        ct = "application/json; charset=utf-8"
        payload = json.dumps(params, ensure_ascii=False)
        canonical_headers = f"content-type:{ct}\nhost:{host}\nx-tc-action:{action.lower()}\n"
        signed_headers = "content-type;host;x-tc-action"
        hashed_request_payload = hashlib.sha256(payload.encode("utf-8")).hexdigest()
        canonical_request = (
            f"{http_request_method}\n{canonical_uri}\n{canonical_querystring}\n"
            f"{canonical_headers}\n{signed_headers}\n{hashed_request_payload}"
        )

        # 2. 拼接待签名串
        algorithm = "TC3-HMAC-SHA256"
        hashed_canonical_request = hashlib.sha256(canonical_request.encode("utf-8")).hexdigest()
        credential_scope = f"{date}/{service}/tc3_request"
        string_to_sign = (
            f"{algorithm}\n{timestamp}\n{credential_scope}\n{hashed_canonical_request}"
        )

        # 3. 计算签名
        secret_date = hmac.new(("TC3" + secret_key).encode("utf-8"), date.encode("utf-8"), hashlib.sha256).digest()
        secret_service = hmac.new(secret_date, service.encode("utf-8"), hashlib.sha256).digest()
        secret_signing = hmac.new(secret_service, "tc3_request".encode("utf-8"), hashlib.sha256).digest()
        signature = hmac.new(secret_signing, string_to_sign.encode("utf-8"), hashlib.sha256).hexdigest()

        # 4. 构建 Authorization
        authorization = (
            f"{algorithm} Credential={secret_id}/{credential_scope}, "
            f"SignedHeaders={signed_headers}, Signature={signature}"
        )

        # 5. 发送请求
        headers = {
            "Authorization": authorization,
            "Content-Type": ct,
            "Host": host,
            "X-TC-Action": action,
            "X-TC-Timestamp": str(timestamp),
            "X-TC-Version": version,
            "X-TC-Region": region,
        }

        resp = requests.post(endpoint, headers=headers, data=payload.encode("utf-8"), timeout=30)
        return resp.json()


# ============================================================
# OCR 服务
# ============================================================

class OCRService:
    """腾讯云 OCR 服务封装"""

    def __init__(self, config):
        self.config = config
        self.signer = TencentCloudSigner()

    def _image_to_base64(self, image_path):
        with open(image_path, "rb") as f:
            return base64.b64encode(f.read()).decode("utf-8")

    def recognize_general_text(self, image_path):
        """通用印刷体识别"""
        img_b64 = self._image_to_base64(image_path)
        params = {"ImageBase64": img_b64}
        result = self.signer.sign(
            self.config["secret_id"],
            self.config["secret_key"],
            "ocr",
            params,
            "GeneralBasicOCR",
            self.config.get("region", "ap-guangzhou"),
        )
        return result

    def recognize_table(self, image_path):
        """表格识别（高精度版）"""
        img_b64 = self._image_to_base64(image_path)
        params = {"ImageBase64": img_b64}
        result = self.signer.sign(
            self.config["secret_id"],
            self.config["secret_key"],
            "ocr",
            params,
            "RecognizeTableAccurateOCR",
            self.config.get("region", "ap-guangzhou"),
            version="2018-11-19",
        )
        return result

    def recognize_formula(self, image_path):
        """数学公式识别"""
        img_b64 = self._image_to_base64(image_path)
        params = {"ImageBase64": img_b64}
        result = self.signer.sign(
            self.config["secret_id"],
            self.config["secret_key"],
            "ocr",
            params,
            "FormulaOCR",
            self.config.get("region", "ap-guangzhou"),
        )
        return result

    def translate_text(self, text, target_lang="en"):
        """文本翻译（腾讯云机器翻译）"""
        params = {
            "SourceText": text,
            "Source": "zh",
            "Target": target_lang,
            "ProjectId": 0,
        }
        result = self.signer.sign(
            self.config["secret_id"],
            self.config["secret_key"],
            "tmt",
            params,
            "TextTranslate",
            self.config.get("region", "ap-guangzhou"),
            version="2018-03-21",
        )
        return result


# ============================================================
# 截图覆盖层
# ============================================================

class SnipOverlay(QWidget):
    """全屏截图选区覆盖层"""
    snipped = pyqtSignal(object)  # 发送截图结果(QImage 或 None)

    def __init__(self):
        super().__init__()
        self.setWindowFlags(Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint | Qt.Tool)
        self.setAttribute(Qt.WA_TranslucentBackground)
        self.setCursor(Qt.CrossCursor)

        # 获取所有屏幕的虚拟几何区域（兼容新旧 PyQt5 版本）
        from PyQt5.QtGui import QGuiApplication
        screens = QGuiApplication.screens()
        if screens:
            from PyQt5.QtCore import QRect
            total = screens[0].virtualGeometry()
            for s in screens[1:]:
                total = total.united(s.geometry())
            self.total_geometry = total
        else:
            from PyQt5.QtCore import QRect
            self.total_geometry = QRect(0, 0, 1920, 1080)

        self.setGeometry(self.total_geometry)

        self.origin = QPoint()
        self.end = QPoint()
        self.dragging = False

        # 截取全屏作为背景
        self.full_screenshot = self._capture_all_screens()

    def _capture_all_screens(self):
        """截取所有屏幕的完整画面"""
        geo = self.total_geometry
        with mss.MSS() as sct:
            monitor = {
                "top": geo.top(),
                "left": geo.left(),
                "width": geo.width(),
                "height": geo.height(),
            }
            shot = sct.grab(monitor)
            img = Image.frombytes("RGB", shot.size, shot.bgra, "raw", "BGRX")
            qimg = QImage(img.tobytes(), img.width, img.height, img.width * 3, QImage.Format_RGB888)
            return QPixmap.fromImage(qimg)

    def paintEvent(self, event):
        painter = QPainter(self)
        # 绘制全屏截图
        painter.drawPixmap(0, 0, self.full_screenshot)

        # 半透明遮罩
        if not self.dragging and self.origin.isNull():
            painter.fillRect(self.rect(), QColor(0, 0, 0, 100))
        else:
            # 选区外半透明
            painter.fillRect(self.rect(), QColor(0, 0, 0, 100))

        if self.dragging and not self.origin.isNull() and not self.end.isNull():
            rect = QRect(self.origin, self.end).normalized()

            # 选区内显示原图
            painter.drawPixmap(rect, self.full_screenshot, rect)

            # 边框
            pen = QPen(QColor(0, 174, 255), 2)
            painter.setPen(pen)
            painter.drawRect(rect)

            # 尺寸标签
            size_text = f"{rect.width()} x {rect.height()}"
            painter.setPen(QColor(255, 255, 255))
            painter.fillRect(rect.x(), rect.y() - 22, 120, 20, QColor(0, 174, 255))
            painter.setFont(QFont("Microsoft YaHei", 9))
            painter.drawText(rect.x() + 5, rect.y() - 7, size_text)

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.origin = event.pos()
            self.end = event.pos()
            self.dragging = True
            self.update()

    def mouseMoveEvent(self, event):
        if self.dragging:
            self.end = event.pos()
            self.update()

    def mouseReleaseEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.dragging = False
            rect = QRect(self.origin, self.end).normalized()

            if rect.width() < 5 or rect.height() < 5:
                self.close()
                self.snipped.emit(None)
                return

            # 从全屏截图中裁剪选区
            cropped = self.full_screenshot.copy(rect)
            self.close()
            self.snipped.emit(cropped)

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Escape:
            self.close()
            self.snipped.emit(None)


# ============================================================
# 主窗口
# ============================================================

STYLE_SHEET = """
QMainWindow { background-color: #f5f6fa; }
QGroupBox { 
    font-weight: bold; font-size: 13px; 
    border: 1px solid #dcdde1; border-radius: 6px; 
    margin-top: 10px; padding-top: 14px;
}
QGroupBox::title { subcontrol-origin: margin; left: 10px; padding: 0 5px; }
QPushButton {
    background-color: #00aaff; color: white; border: none; 
    border-radius: 5px; padding: 8px 16px; font-size: 13px; font-weight: bold;
}
QPushButton:hover { background-color: #0095e0; }
QPushButton:pressed { background-color: #0080c0; }
QPushButton:disabled { background-color: #bdc3c7; }
QPushButton#secondary { background-color: #7f8c8d; }
QPushButton#secondary:hover { background-color: #6c7a7d; }
QTabWidget::pane { border: 1px solid #dcdde1; border-radius: 4px; }
QTabBar::tab { 
    background: #dcdde1; padding: 8px 20px; 
    border-top-left-radius: 4px; border-top-right-radius: 4px;
    font-size: 13px; margin-right: 2px;
}
QTabBar::tab:selected { background: #00aaff; color: white; }
QTableWidget { gridline-color: #dcdde1; font-size: 13px; }
QTableWidget::item { padding: 4px; }
QStatusBar { font-size: 12px; color: #7f8c8d; }
QLabel#title { font-size: 18px; font-weight: bold; color: #2c3e50; }
QLabel#hint { font-size: 12px; color: #95a5a6; }
"""

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.config = load_config()
        self.ocr_service = OCRService(self.config)
        self.current_screenshot = None  # QImage
        self.current_screenshot_path = None
        self.last_text_result = ""
        self.last_table_result = None  # list of list of str
        self.last_formula_result = ""

        self.init_ui()
        self.init_tray()
        self.register_hotkey()

    def init_ui(self):
        self.setWindowTitle("截图OCR工具")
        self.setMinimumSize(900, 650)
        self.setStyleSheet(STYLE_SHEET)

        central = QWidget()
        self.setCentralWidget(central)
        layout = QVBoxLayout(central)
        layout.setSpacing(10)
        layout.setContentsMargins(15, 15, 15, 15)

        # 顶部标题
        title = QLabel("📷 截图OCR工具")
        title.setObjectName("title")
        layout.addWidget(title)

        hint = QLabel("快捷键 Ctrl+Alt+A 截图 | 支持文字、表格、公式识别")
        hint.setObjectName("hint")
        layout.addWidget(hint)

        # 按钮区
        btn_layout = QHBoxLayout()

        self.btn_snip = QPushButton("🖼 截图识别")
        self.btn_snip.clicked.connect(self.start_snip)
        btn_layout.addWidget(self.btn_snip)

        self.btn_reco_text = QPushButton("📝 识别文字")
        self.btn_reco_text.clicked.connect(lambda: self.do_recognize("text"))
        self.btn_reco_text.setEnabled(False)
        btn_layout.addWidget(self.btn_reco_text)

        self.btn_reco_table = QPushButton("📊 识别表格")
        self.btn_reco_table.clicked.connect(lambda: self.do_recognize("table"))
        self.btn_reco_table.setEnabled(False)
        btn_layout.addWidget(self.btn_reco_table)

        self.btn_reco_formula = QPushButton("🔢 识别公式")
        self.btn_reco_formula.clicked.connect(lambda: self.do_recognize("formula"))
        self.btn_reco_formula.setEnabled(False)
        btn_layout.addWidget(self.btn_reco_formula)

        self.btn_config = QPushButton("⚙ 设置")
        self.btn_config.setObjectName("secondary")
        self.btn_config.clicked.connect(self.open_settings)
        btn_layout.addWidget(self.btn_config)

        btn_layout.addStretch()
        layout.addLayout(btn_layout)

        # 截图预览
        preview_group = QGroupBox("截图预览")
        preview_layout = QHBoxLayout(preview_group)
        self.preview_label = QLabel("截图后在此显示")
        self.preview_label.setAlignment(Qt.AlignCenter)
        self.preview_label.setStyleSheet("color: #bdc3c7; font-size: 14px; border: 2px dashed #dcdde1; border-radius: 6px;")
        self.preview_label.setMinimumHeight(180)
        preview_layout.addWidget(self.preview_label)
        layout.addWidget(preview_group)

        # 结果 Tab
        self.tabs = QTabWidget()

        # 文字结果
        self.text_edit = QTextEdit()
        self.text_edit.setFont(QFont("Microsoft YaHei", 11))
        self.text_edit.setPlaceholderText("识别的文字将显示在此...")
        self.tabs.addTab(self.text_edit, "📝 文字")

        # 表格结果
        self.table_widget = QTableWidget()
        self.tabs.addTab(self.table_widget, "📊 表格")

        # 公式结果
        self.formula_edit = QTextEdit()
        self.formula_edit.setFont(QFont("Consolas", 11))
        self.formula_edit.setPlaceholderText("识别的公式 (LaTeX) 将显示在此...")
        self.tabs.addTab(self.formula_edit, "🔢 公式")

        layout.addWidget(self.tabs, stretch=1)

        # 导出按钮
        export_layout = QHBoxLayout()
        export_layout.addWidget(QLabel("导出:"))

        self.btn_export_excel = QPushButton("Excel (.xlsx)")
        self.btn_export_excel.setObjectName("secondary")
        self.btn_export_excel.clicked.connect(lambda: self.export_result("excel"))
        export_layout.addWidget(self.btn_export_excel)

        self.btn_export_csv = QPushButton("CSV")
        self.btn_export_csv.setObjectName("secondary")
        self.btn_export_csv.clicked.connect(lambda: self.export_result("csv"))
        export_layout.addWidget(self.btn_export_csv)

        self.btn_export_md = QPushButton("Markdown")
        self.btn_export_md.setObjectName("secondary")
        self.btn_export_md.clicked.connect(lambda: self.export_result("markdown"))
        export_layout.addWidget(self.btn_export_md)

        self.btn_copy = QPushButton("📋 复制到剪贴板")
        self.btn_copy.setObjectName("secondary")
        self.btn_copy.clicked.connect(self.copy_to_clipboard)
        export_layout.addWidget(self.btn_copy)

        self.btn_translate = QPushButton("🌐 翻译")
        self.btn_translate.setObjectName("secondary")
        self.btn_translate.clicked.connect(self.translate_result)
        export_layout.addWidget(self.btn_translate)

        export_layout.addStretch()
        layout.addLayout(export_layout)

        # 状态栏
        self.status = QStatusBar()
        self.setStatusBar(self.status)
        self.status.showMessage("就绪 | 快捷键 Ctrl+Alt+A 截图")

        # 检查配置
        if not self.config.get("secret_id") or not self.config.get("secret_key"):
            QTimer.singleShot(500, self.prompt_config)

    def init_tray(self):
        """系统托盘"""
        self.tray = QSystemTrayIcon(self)
        self.tray.setIcon(self.style().standardIcon(self.style().SP_ComputerIcon))
        self.tray.setToolTip("截图OCR工具")

        tray_menu = QMenu()
        act_snip = QAction("📷 截图识别", self)
        act_snip.triggered.connect(self.start_snip)
        tray_menu.addAction(act_snip)

        act_show = QAction("显示主窗口", self)
        act_show.triggered.connect(self.show_normal)
        tray_menu.addAction(act_show)

        act_quit = QAction("退出", self)
        act_quit.triggered.connect(QApplication.quit)
        tray_menu.addAction(act_quit)

        self.tray.setContextMenu(tray_menu)
        self.tray.show()
        self.tray.activated.connect(self.on_tray_activated)

    def on_tray_activated(self, reason):
        if reason == QSystemTrayIcon.DoubleClick:
            self.show_normal()

    def show_normal(self):
        self.showNormal()
        self.activateWindow()
        self.raise_()

    def register_hotkey(self):
        """注册全局热键 Ctrl+Alt+A"""
        try:
            import keyboard
            keyboard.add_hotkey("ctrl+alt+a", self.start_snip)
        except ImportError:
            # 如果没有 keyboard 库，使用定时器模拟
            pass

    # --------------------------------------------------------
    # 截图
    # --------------------------------------------------------

    def start_snip(self):
        """启动截图选区"""
        self.showMinimized()
        QApplication.processEvents()
        QTimer.singleShot(300, self._do_snip)

    def _do_snip(self):
        self.overlay = SnipOverlay()
        self.overlay.showFullScreen()
        self.overlay.snipped.connect(self.on_snipped)

    def on_snipped(self, pixmap):
        """截图完成回调"""
        self.show_normal()

        if pixmap is None:
            self.status.showMessage("已取消截图")
            return

        self.current_screenshot = pixmap.toImage()

        # 保存临时文件
        temp_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "temp_screenshot.png")
        pixmap.save(temp_path, "PNG")
        self.current_screenshot_path = temp_path

        # 显示预览
        scaled = pixmap.scaledToWidth(600, Qt.SmoothTransformation)
        if scaled.height() > 180:
            scaled = pixmap.scaledToHeight(180, Qt.SmoothTransformation)
        self.preview_label.setPixmap(scaled)
        self.preview_label.setStyleSheet("")

        # 启用识别按钮
        self.btn_reco_text.setEnabled(True)
        self.btn_reco_table.setEnabled(True)
        self.btn_reco_formula.setEnabled(True)

        self.status.showMessage("截图完成，请选择识别类型")

    # --------------------------------------------------------
    # OCR 识别
    # --------------------------------------------------------

    def do_recognize(self, recog_type):
        if not self.current_screenshot_path:
            QMessageBox.warning(self, "提示", "请先截图")
            return

        if not self.config.get("secret_id") or not self.config.get("secret_key"):
            self.prompt_config()
            return

        self.status.showMessage(f"正在识别{self._recog_name(recog_type)}...")
        QApplication.processEvents()

        try:
            if recog_type == "text":
                result = self.ocr_service.recognize_general_text(self.current_screenshot_path)
                self._handle_text_result(result)
            elif recog_type == "table":
                result = self.ocr_service.recognize_table(self.current_screenshot_path)
                self._handle_table_result(result)
            elif recog_type == "formula":
                result = self.ocr_service.recognize_formula(self.current_screenshot_path)
                self._handle_formula_result(result)
        except Exception as e:
            QMessageBox.critical(self, "识别失败", f"API调用出错:\n{str(e)}")
            self.status.showMessage("识别失败")

    def _recog_name(self, t):
        return {"text": "文字", "table": "表格", "formula": "公式"}.get(t, "")

    def _handle_text_result(self, result):
        """处理文字识别结果"""
        if "Response" not in result:
            QMessageBox.warning(self, "错误", f"返回格式异常:\n{json.dumps(result, ensure_ascii=False)}")
            return

        resp = result["Response"]
        if "Error" in resp:
            QMessageBox.critical(self, "识别失败", f"错误码: {resp['Error']['Code']}\n{resp['Error']['Message']}")
            return

        texts = []
        if "TextDetections" in resp:
            for item in resp["TextDetections"]:
                texts.append(item.get("DetectedText", ""))

        full_text = "\n".join(texts)
        self.last_text_result = full_text
        self.text_edit.setPlainText(full_text)
        self.tabs.setCurrentIndex(0)
        self.status.showMessage(f"文字识别完成，共 {len(texts)} 段")

    def _handle_table_result(self, result):
        """处理表格识别结果"""
        if "Response" not in result:
            QMessageBox.warning(self, "错误", f"返回格式异常:\n{json.dumps(result, ensure_ascii=False)}")
            return

        resp = result["Response"]
        if "Error" in resp:
            QMessageBox.critical(self, "识别失败", f"错误码: {resp['Error']['Code']}\n{resp['Error']['Message']}")
            return

        tables = resp.get("TableDetections", [])
        if not tables:
            self.status.showMessage("未识别到表格")
            return

        # 取第一个表格
        table = tables[0]
        table_body = table.get("TableBody", "")

        # 解析 HTML 表格
        rows = self._parse_html_table(table_body)
        self.last_table_result = rows

        # 显示到 QTableWidget
        self._display_table(rows)
        self.tabs.setCurrentIndex(1)
        self.status.showMessage(f"表格识别完成，{len(rows)} 行 x {len(rows[0]) if rows else 0} 列")

    def _parse_html_table(self, html):
        """简单解析 HTML 表格为二维列表"""
        rows = []
        current_row = []
        current_cell = []

        i = 0
        in_cell = False
        in_row = False

        while i < len(html):
            if html[i:i+3].lower() == "<tr":
                in_row = True
                current_row = []
                # 跳到 >
                gt = html.find(">", i)
                if gt == -1:
                    break
                i = gt + 1
                continue
            elif html[i:i+5].lower() == "</tr>":
                if current_row:
                    rows.append(current_row)
                current_row = []
                in_row = False
                i += 5
                continue
            elif html[i:i+3].lower() in ("<td", "<th"):
                in_cell = True
                current_cell = []
                gt = html.find(">", i)
                if gt == -1:
                    break
                i = gt + 1
                continue
            elif html[i:i+5].lower() == "</td>" or html[i:i+5].lower() == "</th>":
                if in_cell:
                    cell_text = "".join(current_cell).strip()
                    current_row.append(cell_text)
                in_cell = False
                i += 5
                continue
            else:
                if in_cell:
                    current_cell.append(html[i])
                i += 1

        return rows

    def _display_table(self, rows):
        """在 QTableWidget 中显示表格"""
        if not rows:
            return

        max_cols = max(len(r) for r in rows)
        self.table_widget.setRowCount(len(rows))
        self.table_widget.setColumnCount(max_cols)

        for r, row in enumerate(rows):
            for c, cell in enumerate(row):
                item = QTableWidgetItem(cell)
                self.table_widget.setItem(r, c, item)

        self.table_widget.resizeColumnsToContents()
        self.table_widget.resizeRowsToContents()

    def _handle_formula_result(self, result):
        """处理公式识别结果"""
        if "Response" not in result:
            QMessageBox.warning(self, "错误", f"返回格式异常:\n{json.dumps(result, ensure_ascii=False)}")
            return

        resp = result["Response"]
        if "Error" in resp:
            QMessageBox.critical(self, "识别失败", f"错误码: {resp['Error']['Code']}\n{resp['Error']['Message']}")
            return

        formulas = []
        if "Formulas" in resp:
            for item in resp["Formulas"]:
                formulas.append(item.get("DetectedFormula", ""))

        formula_text = "\n\n".join(formulas)
        self.last_formula_result = formula_text
        self.formula_edit.setPlainText(formula_text)
        self.tabs.setCurrentIndex(2)
        self.status.showMessage(f"公式识别完成，共 {len(formulas)} 个公式")

    # --------------------------------------------------------
    # 导出
    # --------------------------------------------------------

    def export_result(self, fmt):
        """导出结果"""
        current_tab = self.tabs.currentIndex()

        file_filter = {
            "excel": "Excel 文件 (*.xlsx)",
            "csv": "CSV 文件 (*.csv)",
            "markdown": "Markdown 文件 (*.md)",
        }

        default_name = {
            "excel": "ocr_result.xlsx",
            "csv": "ocr_result.csv",
            "markdown": "ocr_result.md",
        }

        path, _ = QFileDialog.getSaveFileName(self, "保存文件", default_name[fmt], file_filter[fmt])
        if not path:
            return

        try:
            if fmt == "excel":
                self._export_excel(path, current_tab)
            elif fmt == "csv":
                self._export_csv(path, current_tab)
            elif fmt == "markdown":
                self._export_markdown(path, current_tab)

            self.status.showMessage(f"已导出到 {path}")
            QMessageBox.information(self, "导出成功", f"文件已保存到:\n{path}")

        except Exception as e:
            QMessageBox.critical(self, "导出失败", str(e))

    def _get_current_content(self, tab_index):
        """获取当前 Tab 的内容"""
        if tab_index == 0:  # 文字
            return ("text", self.text_edit.toPlainText())
        elif tab_index == 1:  # 表格
            return ("table", self.last_table_result)
        elif tab_index == 2:  # 公式
            return ("text", self.formula_edit.toPlainText())
        return ("text", "")

    def _export_excel(self, path, tab_index):
        import openpyxl
        from openpyxl.styles import Font, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "OCR结果"

        content_type, content = self._get_current_content(tab_index)

        if content_type == "table" and content:
            for r, row in enumerate(content, 1):
                for c, cell in enumerate(row, 1):
                    cell_obj = ws.cell(row=r, column=c, value=cell)
                    cell_obj.alignment = Alignment(vertical="center", wrap_text=True)
            # 自适应列宽
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        max_len = max(max_len, len(str(cell.value or "")))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = min(max_len + 4, 50)
        else:
            # 纯文本写入
            lines = content.split("\n") if content else [""]
            for r, line in enumerate(lines, 1):
                ws.cell(row=r, column=1, value=line)

        wb.save(path)

    def _export_csv(self, path, tab_index):
        content_type, content = self._get_current_content(tab_index)

        # Windows Excel 用 UTF-8 BOM 打开 CSV
        with open(path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            if content_type == "table" and content:
                for row in content:
                    writer.writerow(row)
            else:
                for line in (content or "").split("\n"):
                    writer.writerow([line])

    def _export_markdown(self, path, tab_index):
        content_type, content = self._get_current_content(tab_index)

        with open(path, "w", encoding="utf-8") as f:
            if content_type == "table" and content:
                # Markdown 表格
                if content:
                    max_cols = max(len(r) for r in content)
                    # 补齐
                    for row in content:
                        while len(row) < max_cols:
                            row.append("")

                    # 表头
                    f.write("| " + " | ".join(content[0]) + " |\n")
                    f.write("| " + " | ".join(["---"] * max_cols) + " |\n")
                    for row in content[1:]:
                        f.write("| " + " | ".join(row) + " |\n")
            else:
                f.write(content or "")

    # --------------------------------------------------------
    # 复制 / 翻译
    # --------------------------------------------------------

    def copy_to_clipboard(self):
        tab_index = self.tabs.currentIndex()
        content_type, content = self._get_current_content(tab_index)

        if content_type == "table" and content:
            # 表格复制为 TSV
            text = "\n".join("\t".join(row) for row in content)
        else:
            text = content or ""

        clipboard = QApplication.clipboard()
        clipboard.setText(text)
        self.status.showMessage("已复制到剪贴板")

    def translate_result(self):
        """翻译当前 Tab 的文字"""
        tab_index = self.tabs.currentIndex()
        content_type, content = self._get_current_content(tab_index)

        if content_type == "table":
            text = "\n".join("\t".join(row) for row in (content or []))
        else:
            text = content or ""

        if not text.strip():
            QMessageBox.warning(self, "提示", "没有可翻译的内容")
            return

        if not self.config.get("secret_id") or not self.config.get("secret_key"):
            self.prompt_config()
            return

        self.status.showMessage("正在翻译...")
        QApplication.processEvents()

        try:
            target = self.config.get("translate_lang", "en")
            result = self.ocr_service.translate_text(text, target)

            if "Response" not in result:
                QMessageBox.warning(self, "错误", f"翻译返回异常:\n{json.dumps(result, ensure_ascii=False)}")
                return

            resp = result["Response"]
            if "Error" in resp:
                QMessageBox.critical(self, "翻译失败", f"错误码: {resp['Error']['Code']}\n{resp['Error']['Message']}")
                return

            translated = resp.get("TargetText", "")
            # 在原文后追加翻译
            if content_type == "table":
                self.text_edit.setPlainText(f"【原文】\n{text}\n\n【翻译】\n{translated}")
                self.tabs.setCurrentIndex(0)
            else:
                current_text = self.text_edit.toPlainText() if tab_index == 0 else self.formula_edit.toPlainText()
                self.text_edit.setPlainText(f"{current_text}\n\n--- 翻译 ---\n{translated}")
                self.tabs.setCurrentIndex(0)

            self.status.showMessage("翻译完成")

        except Exception as e:
            QMessageBox.critical(self, "翻译失败", str(e))

    # --------------------------------------------------------
    # 设置
    # --------------------------------------------------------

    def prompt_config(self):
        QMessageBox.information(
            self, "需要配置",
            "使用前需要配置腾讯云 API 密钥。\n\n"
            "获取方式：\n"
            "1. 访问 https://console.cloud.tencent.com/cam/capi\n"
            "2. 创建或查看 SecretId 和 SecretKey\n"
            "3. 确保已开通 OCR 和机器翻译服务\n\n"
            "点击确定进入设置页面。"
        )
        self.open_settings()

    def open_settings(self):
        """打开设置对话框"""
        from PyQt5.QtWidgets import QDialog, QFormLayout, QLineEdit, QComboBox, QDialogButtonBox

        dlg = QDialog(self)
        dlg.setWindowTitle("设置 - 腾讯云API")
        dlg.setMinimumWidth(450)

        layout = QFormLayout(dlg)

        id_edit = QLineEdit(self.config.get("secret_id", ""))
        id_edit.setPlaceholderText("SecretId")

        key_edit = QLineEdit(self.config.get("secret_key", ""))
        key_edit.setEchoMode(QLineEdit.Password)
        key_edit.setPlaceholderText("SecretKey")

        region_combo = QComboBox()
        regions = ["ap-guangzhou", "ap-beijing", "ap-shanghai", "ap-chengdu", "ap-chongqing", "ap-nanjing"]
        region_combo.addItems(regions)
        region_combo.setCurrentText(self.config.get("region", "ap-guangzhou"))

        translate_combo = QComboBox()
        translate_combo.addItems(["en (英语)", "ja (日语)", "ko (韩语)", "fr (法语)", "de (德语)", "ru (俄语)", "es (西班牙语)"])
        lang_map = {"en": "en (英语)", "ja": "ja (日语)", "ko": "ko (韩语)", "fr": "fr (法语)", "de": "de (德语)", "ru": "ru (俄语)", "es": "es (西班牙语)"}
        translate_combo.setCurrentText(lang_map.get(self.config.get("translate_lang", "en"), "en (英语)"))

        layout.addRow("SecretId:", id_edit)
        layout.addRow("SecretKey:", key_edit)
        layout.addRow("区域:", region_combo)
        layout.addRow("翻译目标语言:", translate_combo)

        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(dlg.accept)
        btns.rejected.connect(dlg.reject)
        layout.addRow(btns)

        if dlg.exec_():
            self.config["secret_id"] = id_edit.text().strip()
            self.config["secret_key"] = key_edit.text().strip()
            self.config["region"] = region_combo.currentText()
            # 解析翻译语言
            tl = translate_combo.currentText().split(" ")[0]
            self.config["translate_lang"] = tl
            save_config(self.config)

            # 重新初始化 OCR 服务
            self.ocr_service = OCRService(self.config)

            self.status.showMessage("设置已保存")

    # --------------------------------------------------------
    # 窗口事件
    # --------------------------------------------------------

    def closeEvent(self, event):
        """关闭时最小化到托盘"""
        event.ignore()
        self.hide()
        self.tray.showMessage("截图OCR工具", "程序已最小化到托盘，双击图标恢复", QSystemTrayIcon.Information, 2000)


# ============================================================
# 入口
# ============================================================

def main():
    # 高 DPI 支持
    QApplication.setAttribute(Qt.AA_EnableHighDpiScaling, True)
    QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps, True)

    app = QApplication(sys.argv)
    app.setQuitOnLastWindowClosed(False)

    window = MainWindow()
    window.show()

    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
